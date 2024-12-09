import openpyxl

wb = openpyxl.load_workbook('videogamesales.xlsx')
ws = wb.active

ws = wb['vgsales']

row_position = 2
col_position = 7

total_sales = ((ws.cell(row=row_position, column=col_position).value)
               + (ws.cell(row=row_position, column=col_position + 1).value)
               + (ws.cell(row=row_position, column=col_position + 2).value)
               + (ws.cell(row=row_position, column=col_position + 3).value))

ws.cell(row=2, column=11).value = total_sales
wb.save('videogamesales.xlsx')
