import openpyxl

wb = openpyxl.load_workbook('../data/videogamesales.xlsx')
ws = wb.active

# ws = wb['vgsales']

row_position = 1
for i in range(1, ws.max_row):
    row_position += 1  # i = i + 1
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_sales = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_sales

new_row = (1, "The Legend of Zelda", 1986, "Action", "Nintendo", 3.74, 0.93, 1.69, 0.14, 6.51, 6.5)
ws.append(new_row)

# wb.save('videogamesales.xlsx')  # save the updated workbook)
# wb.save('../data/videogamesales.xlsx')  # save the updated workbook)

values = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
print(values)
# [1, 'The Legend of Zelda', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51, 6.5]

ws.delete_rows(ws.max_row, 1)
# wb.save('videogamesales.xlsx')  # save the updated workbook)
# wb.save('../data/videogamesales.xlsx')

ws['P1'] = "Average Sales"
ws['P2'] = '= AVERAGE(K2:K16220)'

# wb.save('../data/videogamesales.xlsx')

ws['Q1'] = "Number of Populated Cells"
ws['Q2'] = "=COUNTA(E2:E16220)"

# wb.save('../data/videogamesales.xlsx')

ws['R1'] = "Number of Rows with Sports Genre"
ws['R2'] = '=COUNTIF(E2:E16220, "Sports")'

# wb.save('../data/videogamesales.xlsx')

ws['S1'] = "Total sports Sales"
ws['S2'] = '=SUMIF(E2:E16220, "Sports",K2:K16220)'

ws['T1'] = "Rounded sum of Sports Sales"
ws['T2'] = '=CEILING(S2,25)'

print(ws.title)  # vgsales

ws.title = "Video Games Sales Data"

print(wb.sheetnames)

wb.create_sheet('Empty Sheet ')
print(wb.sheetnames)

wb.save('../data/videogamesales.xlsx')

wb.remove(wb['Empty Sheet '])
print(wb.sheetnames)

wb.save('../data/videogamesales.xlsx')

wb.copy_worksheet(wb['Video Games Sales Data'])
wb.save('../data/videogamesales.xlsx')
wb.save('../data/vgsales.xlsx')
