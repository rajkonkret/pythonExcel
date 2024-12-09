import openpyxl

wb = openpyxl.load_workbook('videogamesales.xlsx')
ws = wb.active

ws = wb['vgsales']

print('Total number of rows: ' + str(ws.max_row) + " And total number of columns: " + str(ws.max_column))
print("Value in cell A1 is: ", ws['A1'].value)  # Value in cell A1 is:  Rank

values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column)]
print(values)

data = [ws.cell(row=i, column=2).value for i in range(2, 12)]
print(data)

my_list = list()  # [] pusta lista

for value in ws.iter_rows(
        min_row=1, max_row=11, min_col=1, max_col=6,
        values_only=True):
    my_list.append(value)

for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
    print("{:.<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1, ele2, ele3, ele4, ele5, ele6))

ws['K1'] = "Sum of Sales"

wb.save('videogamesales.xlsx')