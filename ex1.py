# import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# pip install openpyxl

wb = Workbook()
ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])

wb.save('sample.xlsx')
wb.close()

print("Excel file has been created successfully.")
print("File name: sample.xlsx")
print("Location: current directory")
print("Open this file in Excel to see the results.")

workbook = load_workbook('sample.xlsx')
sheet = workbook.active
print(sheet)

print(sheet['A1'].value)  # 42
for row in sheet.iter_rows(min_row=1, max_row=3):
    for cell in row:
        print(cell.value)
